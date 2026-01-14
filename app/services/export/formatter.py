from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


class ExportFormatter:
    """Handles Excel styling and worksheet-level formatting."""
    @staticmethod
    def apply_header_style(ws, row=1):
        for cell in ws[row]:
            cell.font = Font(bold=True)

    @staticmethod
    def auto_fit_columns(ws, min_width=15):
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        lines = str(cell.value).split('\n')
                        max_line_length = max(len(line) for line in lines)
                        if max_line_length > max_length:
                            max_length = max_line_length
                except:
                    pass
            adjusted_width = max(max_length + 2, min_width)
            ws.column_dimensions[column_letter].width = min(adjusted_width, 50)

    @staticmethod
    def apply_wrap_text(cell, vertical='top'):
        cell.alignment = Alignment(wrap_text=True, vertical=vertical)
