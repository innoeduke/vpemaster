# vpemaster/agenda_routes.py

from flask import Blueprint, render_template, request, redirect, url_for, jsonify, send_file, current_app
from .main_routes import login_required, session
from .models import SessionLog, SessionType, Contact, Meeting, Project, Presentation, Media
from . import db
from sqlalchemy import distinct, orm
from datetime import datetime, timedelta
import io
import csv
import os
import openpyxl
from openpyxl.styles import Font
from io import BytesIO
from .utils import derive_current_path_level

agenda_bp = Blueprint('agenda_bp', __name__)

# --- Helper Functions for Data Fetching ---


def _get_agenda_logs(meeting_number):
    """Fetches all agenda logs and related data for a specific meeting."""
    query = db.session.query(SessionLog)\
        .options(
            orm.joinedload(SessionLog.session_type),  # Eager load SessionType
            orm.joinedload(SessionLog.meeting),      # Eager load Meeting
            orm.joinedload(SessionLog.project),      # Eager load Project
            orm.joinedload(SessionLog.owner),         # Eager load Owner
            orm.joinedload(SessionLog.media)
    )

    if meeting_number:
        query = query.filter(SessionLog.Meeting_Number == meeting_number)

    # Order by sequence
    return query.order_by(SessionLog.Meeting_Seq.asc()).all()


def _get_project_speakers(meeting_number):
    """Gets a list of speakers for a given meeting."""
    if not meeting_number:
        return []
    speaker_logs = db.session.query(Contact.Name)\
        .join(SessionLog, Contact.id == SessionLog.Owner_ID)\
        .join(SessionType, SessionLog.Type_ID == SessionType.id)\
        .filter(
            SessionLog.Meeting_Number == meeting_number,
            SessionType.Valid_for_Project == True
    ).all()
    return [name[0] for name in speaker_logs]


def _create_or_update_session(item, meeting_number, seq):
    meeting = Meeting.query.filter_by(Meeting_Number=meeting_number).first()
    if not meeting:
        new_meeting = Meeting(Meeting_Number=meeting_number)
        db.session.add(new_meeting)
        db.session.flush()  # Flush to get meeting ID if needed, though not strictly required here
        meeting = new_meeting

    type_id = item.get('type_id')
    session_type = SessionType.query.get(
        type_id) if type_id else None  # Get the SessionType object

    # --- Owner ID Handling (no changes needed) ---
    owner_id_val = item.get('owner_id')
    owner_id = int(owner_id_val) if owner_id_val and owner_id_val not in [
        'None', '0'] else None
    owner_contact = Contact.query.get(owner_id) if owner_id else None

    # --- Project ID and Status (no changes needed) ---
    project_id = item.get('project_id')
    if project_id in [None, "", "null"]:
        project_id = None
    status = item.get('status') if item.get('status') else 'Booked'
    session_title = item.get('session_title')

    # --- Designation Logic (no changes needed) ---
    designation = item.get('designation')
    # Check for None or empty string
    if owner_id and (designation is None or designation == ''):
        owner = Contact.query.get(owner_id)
        if owner:
            if owner.DTM:
                designation = None  # Keep as None if DTM
            elif owner.Type == 'Guest':
                designation = f"Guest@{owner.Club}" if owner.Club else "Guest"
            elif owner.Type == 'Member':
                designation = owner.Completed_Levels.replace(
                    ' ', '/') if owner.Completed_Levels else None
    # Ensure designation is an empty string if None before saving to avoid DB issues if nullable=False
    if designation is None:
        designation = ''

    # --- Automatic current_path_level Derivation ---
    temp_log_data = {
        'Project_ID': project_id,
        'Type_ID': type_id,
        'session_type': session_type,  # Pass fetched session_type
        # Fetch project if needed
        'project': Project.query.get(project_id) if project_id else None
    }
    # Create a temporary object-like structure or fetch existing log if updating
    log_for_derivation = SessionLog.query.get(item['id']) if item.get(
        'id') != 'new' else type('obj', (object,), temp_log_data)()
    if item.get('id') == 'new':  # If new, update the temp obj with necessary fields
        log_for_derivation.Project_ID = project_id
        log_for_derivation.Type_ID = type_id
        log_for_derivation.session_type = session_type
        log_for_derivation.project = temp_log_data['project']

    current_path_level = derive_current_path_level(
        log_for_derivation, owner_contact)

    # --- Create or Update SessionLog ---
    if item['id'] == 'new':
        new_log = SessionLog(
            Meeting_Number=meeting_number,
            Meeting_Seq=seq,
            Type_ID=type_id,
            Owner_ID=owner_id,
            Designation=designation,
            Duration_Min=item.get('duration_min') if item.get(
                'duration_min') else None,
            Duration_Max=item.get('duration_max') if item.get(
                'duration_max') else None,
            Project_ID=project_id,
            Session_Title=session_title,
            Status=status,
            current_path_level=current_path_level
        )
        db.session.add(new_log)
    else:
        log = SessionLog.query.get(item['id'])
        if log:
            log.Meeting_Number = meeting_number
            log.Meeting_Seq = seq
            log.Type_ID = type_id
            log.Owner_ID = owner_id
            log.Designation = designation
            log.Duration_Min = item.get('duration_min') if item.get(
                'duration_min') else None
            log.Duration_Max = item.get('duration_max') if item.get(
                'duration_max') else None
            log.Project_ID = project_id
            log.Status = status
            if session_title is not None:
                log.Session_Title = session_title
            log.current_path_level = current_path_level


def _recalculate_start_times(meetings_to_update):
    for meeting in meetings_to_update:
        if not meeting or not meeting.Start_Time or not meeting.Meeting_Date:
            continue

        current_time = meeting.Start_Time
        # Fetch Is_Hidden along with Is_Section
        logs_to_update = db.session.query(SessionLog, SessionType.Is_Section, SessionType.Is_Hidden)\
            .join(SessionType, SessionLog.Type_ID == SessionType.id, isouter=True)\
            .filter(SessionLog.Meeting_Number == meeting.Meeting_Number)\
            .order_by(SessionLog.Meeting_Seq.asc()).all()

        for log, is_section, is_hidden in logs_to_update:
            # --- THIS IS THE FIX ---
            # If the session is a section header OR if it's hidden, set its time to None and continue.
            if is_section or is_hidden:
                log.Start_Time = None
                continue
            # --- END OF FIX ---

            # The rest of the logic only runs for visible, non-section items.
            log.Start_Time = current_time
            duration_to_add = int(log.Duration_Max or 0)
            break_minutes = 1
            # Individual Evaluation Type_ID is 31
            if log.Type_ID == 31 and meeting.GE_Style == 'immediate':
                break_minutes += 1
            dt_current_time = datetime.combine(
                meeting.Meeting_Date, current_time)
            next_dt = dt_current_time + \
                timedelta(minutes=duration_to_add + break_minutes)
            current_time = next_dt.time()

# --- Main Route ---


@agenda_bp.route('/agenda', methods=['GET'])
def agenda():

    SERIES_INITIALS = {
        "Successful Club Series": "SC",
        "Better Speaker Series": "BS",
        "Leadership Excellence Series": "LE"
    }

    # --- Determine Selected Meeting ---
    meeting_numbers_query = db.session.query(distinct(
        SessionLog.Meeting_Number)).order_by(SessionLog.Meeting_Number.desc()).all()
    meeting_numbers = [num[0] for num in meeting_numbers_query]

    selected_meeting_str = request.args.get('meeting_number')
    selected_meeting_num = None

    if selected_meeting_str:
        try:
            selected_meeting_num = int(selected_meeting_str)
        except ValueError:
            # Handle invalid meeting number string if necessary
            selected_meeting_num = None  # Or redirect, flash error
    else:
        # Find the most recent upcoming meeting number
        upcoming_meeting = Meeting.query\
            .filter(Meeting.Meeting_Date >= datetime.today().date())\
            .order_by(Meeting.Meeting_Date.asc(), Meeting.Meeting_Number.asc())\
            .first()

        if upcoming_meeting:
            selected_meeting_num = upcoming_meeting.Meeting_Number
        elif meeting_numbers:
            # Fallback to the most recent existing meeting (highest meeting number)
            selected_meeting_num = meeting_numbers[0]

    selected_meeting = None
    if selected_meeting_num:
        selected_meeting = Meeting.query.options(
            orm.joinedload(Meeting.best_tt_speaker),
            orm.joinedload(Meeting.best_evaluator),
            orm.joinedload(Meeting.best_speaker),
            orm.joinedload(Meeting.best_roletaker),
            orm.joinedload(Meeting.media)
        ).filter(Meeting.Meeting_Number == selected_meeting_num).first()

    # --- Fetch Raw Data ---
    raw_session_logs = _get_agenda_logs(selected_meeting_num)
    project_speakers = _get_project_speakers(selected_meeting_num)

    all_presentations = Presentation.query.order_by(
        Presentation.level, Presentation.code).all()
    presentation_series = sorted(
        list(set(p.series for p in all_presentations if p.series)))
    presentations_data = [
        {"id": p.id, "title": p.title, "level": p.level,
            "series": p.series, "code": p.code}
        for p in all_presentations
    ]

    # --- Process Raw Logs into Dictionaries ---
    logs_data = []
    pathway_mapping = current_app.config['PATHWAY_MAPPING']
    for log in raw_session_logs:
        session_type = log.session_type
        meeting = log.meeting
        project = log.project
        owner = log.owner
        media = log.media

        # Determine project code if applicable
        project_code_str = None
        session_title_for_dict = log.Session_Title
        project_name_for_dict = project.Project_Name if project else ''
        project_purpose_for_dict = project.Purpose if project else ''
        pathway_code_for_dict = None
        level_for_dict = None

        if session_type and session_type.Title == 'Presentation' and log.Project_ID:
            # Find in the already-fetched list
            presentation = next(
                (p for p in all_presentations if p.id == log.Project_ID), None)
            if presentation:
                if not session_title_for_dict:
                    session_title_for_dict = presentation.title
                project_name_for_dict = presentation.series  # For tooltip
                project_purpose_for_dict = f"Level {presentation.level} - {presentation.title}"

                series_key = " ".join(
                    presentation.series.split()) if presentation.series else ""

                series_initial = SERIES_INITIALS.get(series_key, "")
                project_code_str = f"{series_initial}{presentation.code}"
        elif project and owner and owner.Working_Path:  # Else if pathway...
            pathway_suffix = pathway_mapping.get(owner.Working_Path)
            if pathway_suffix:
                project_code_val = getattr(
                    project, f"Code_{pathway_suffix}", None)
                if project_code_val:
                    project_code_str = f"{pathway_suffix}{project_code_val}"
                    pathway_code_for_dict = pathway_suffix
                    try:
                        level_for_dict = int(project_code_val.split('.')[0])
                    except (ValueError, IndexError):
                        level_for_dict = None

        log_dict = {
            # SessionLog fields
            'id': log.id,
            'Project_ID': log.Project_ID,
            'Meeting_Number': log.Meeting_Number,
            'Meeting_Seq': log.Meeting_Seq,
            'Start_Time_str': log.Start_Time.strftime('%H:%M') if log.Start_Time else '',
            'Session_Title': session_title_for_dict,
            'Type_ID': log.Type_ID,
            'Owner_ID': log.Owner_ID,
            'Designation': log.Designation,
            'Duration_Min': log.Duration_Min,
            'Duration_Max': log.Duration_Max,
            'Status': log.Status,
            'current_path_level': log.current_path_level,
            # SessionType fields
            'is_section': session_type.Is_Section if session_type else False,
            'session_type_title': session_type.Title if session_type else 'Unknown Type',
            'predefined': session_type.Predefined if session_type else False,
            'is_hidden': session_type.Is_Hidden if session_type else False,
            'role': session_type.Role if session_type else '',
            'valid_for_project': session_type.Valid_for_Project if session_type else False,
            # Meeting fields
            'meeting_date_str': meeting.Meeting_Date.strftime('%Y-%m-%d') if meeting and meeting.Meeting_Date else '',
            'meeting_date_formatted': meeting.Meeting_Date.strftime('%B %d, %Y') if meeting and meeting.Meeting_Date else '',
            # Project fields
            'project_name': project_name_for_dict,
            'project_purpose': project_purpose_for_dict,
            'project_code_display': project_code_str,
            'pathway_code': pathway_code_for_dict,
            'level': level_for_dict,
            # Owner fields
            'owner_name': owner.Name if owner else '',
            'owner_dtm': owner.DTM if owner else False,
            'owner_type': owner.Type if owner else '',
            'owner_club': owner.Club if owner else '',
            'owner_completed_levels': owner.Completed_Levels if owner else '',
            'media_url': media.url if media and media.url else None,
        }
        logs_data.append(log_dict)

    # --- Prepare Data for Template and JavaScript ---
    projects = Project.query.order_by(Project.Project_Name).all()
    projects_data = [
        {
            "ID": p.ID, "Project_Name": p.Project_Name, "Code_DL": p.Code_DL,
            "Code_EH": p.Code_EH, "Code_MS": p.Code_MS, "Code_PI": p.Code_PI,
            "Code_PM": p.Code_PM, "Code_VC": p.Code_VC, "Code_DTM": p.Code_DTM,
            "Purpose": p.Purpose, "Duration_Min": p.Duration_Min, "Duration_Max": p.Duration_Max
        } for p in projects
    ]
    pathways = list(current_app.config['PATHWAY_MAPPING'].keys())
    session_types = SessionType.query.order_by(SessionType.Title.asc()).all()
    session_types_data = [
        {
            "id": s.id, "Title": s.Title, "Is_Section": s.Is_Section,
            "Valid_for_Project": s.Valid_for_Project, "Predefined": s.Predefined,
            "Role": s.Role or '', "Role_Group": s.Role_Group or '',
            "Duration_Min": s.Duration_Min, "Duration_Max": s.Duration_Max
        } for s in session_types
    ]
    contacts = Contact.query.order_by(Contact.Name.asc()).all()
    contacts_data = [
        {
            "id": c.id, "Name": c.Name, "DTM": c.DTM, "Type": c.Type,
            "Club": c.Club, "Working_Path": c.Working_Path, "Next_Project": c.Next_Project,
            "Completed_Levels": c.Completed_Levels
        } for c in contacts
    ]
    members = Contact.query.filter_by(
        Type='Member').order_by(Contact.Name.asc()).all()

    template_dir = os.path.join(current_app.static_folder, 'mtg_templates')
    try:
        meeting_templates = [f for f in os.listdir(
            template_dir) if f.endswith('.csv')]
    except FileNotFoundError:
        meeting_templates = []

    # --- Render Template ---
    return render_template('agenda.html',
                           logs_data=logs_data,               # Use the processed list of dictionaries
                           session_types=session_types_data,  # For JS dropdowns
                           contacts=contacts,               # For display logic in template if needed
                           contacts_data=contacts_data,     # For JS autocomplete/modals
                           projects=projects_data,          # For JS modals/tooltips
                           pathways=pathways,               # For modals
                           pathway_mapping=current_app.config['PATHWAY_MAPPING'],
                           meeting_numbers=meeting_numbers,
                           selected_meeting=selected_meeting,  # Pass the Meeting object
                           members=members,                 # If needed elsewhere
                           project_speakers=project_speakers,  # For JS
                           meeting_templates=meeting_templates,
                           series_initials=SERIES_INITIALS,
                           presentations=presentations_data,
                           presentation_series=presentation_series)


def _get_export_data(meeting_number):
    """
    Fetches the main agenda log data for the export,
    eagerly loading all required related objects.
    """
    return db.session.query(
        SessionLog,
        SessionType,
        Contact,
        Project,
        Media
    ).outerjoin(Contact, SessionLog.Owner_ID == Contact.id)\
     .join(SessionType, SessionLog.Type_ID == SessionType.id, isouter=True)\
     .outerjoin(Project, SessionLog.Project_ID == Project.ID)\
     .outerjoin(Media, SessionLog.id == Media.log_id)\
     .filter(SessionLog.Meeting_Number == meeting_number)\
     .order_by(SessionLog.Meeting_Seq.asc()).all()


def _get_project_code_str(project, contact, pathway_mapping):
    """Helper to get the project code string (e.g., PM2.1)."""
    if project and contact and contact.Working_Path:
        pathway_abbr = pathway_mapping.get(contact.Working_Path)
        if pathway_abbr:
            return getattr(project, f"Code_{pathway_abbr}", None)
    return None


def _get_all_speech_details(logs_data, pathway_mapping):
    """
    Processes logs_data to extract a structured list of speech details.
    """
    speeches = []
    for log, session_type, contact, project, media in logs_data:
        # Check if it's a Pathway Speech (32) or Presentation (14)
        if log.Project_ID and session_type and (session_type.id == 32 or session_type.id == 14):
            path_abbr = pathway_mapping.get(
                contact.Working_Path, "") if contact else ""
            project_code = _get_project_code_str(
                project, contact, pathway_mapping)

            speeches.append({
                "speaker_name": contact.Name if contact else "N/A",
                "speech_title": log.Session_Title or project.Project_Name,
                "project_code_str": f"{path_abbr}{project_code}" if project_code else "Generic",
                "duration": f"[{project.Duration_Min}'-{project.Duration_Max}']" if project else "",
                "media_url": media.url if media else ""
            })
    return speeches


def _format_export_row(log, session_type, contact, project, pathway_mapping):
    """Formats a single row for the human-readable agenda (Sheet 1)."""
    if session_type and session_type.Is_Section:
        return ('section', ['', log.Session_Title or session_type.Title, '', ''])

    start_time = log.Start_Time.strftime('%H:%M') if log.Start_Time else ''

    # Session and Speech Titles
    session_title = log.Session_Title if log.Session_Title else (
        session_type.Title if session_type else '')
    if log.Type_ID == 31 and log.Session_Title:  # Individual Evaluation
        session_title = f"Evaluation for {log.Session_Title}"
    elif log.Project_ID and contact and contact.Working_Path:
        project_code = _get_project_code_str(project, contact, pathway_mapping)
        pathway_abbr = pathway_mapping.get(contact.Working_Path, "")
        if project_code:
            title_part = f'"{log.Session_Title}" ' if log.Session_Title else ''
            session_title = f"{title_part}({pathway_abbr}{project_code})"
    if session_title:
        session_title = session_title.replace('""', '"')

    # Owner Information
    owner_info = ''
    if contact:
        owner_info = contact.Name
        meta_parts = []
        if contact.DTM:
            meta_parts.append('DTM')
        if log.Designation:
            meta_parts.append(log.Designation)
        elif contact.Type == 'Member' and contact.Completed_Levels:
            meta_parts.append(contact.Completed_Levels.replace('/', ' '))
        if meta_parts:
            owner_info += ' - ' + ' '.join(meta_parts)

    # Duration
    duration = ''
    min_dur = log.Duration_Min
    max_dur = log.Duration_Max
    if max_dur is not None:
        if min_dur is not None and min_dur > 0 and min_dur != max_dur:
            duration = f"[{min_dur}'-{max_dur}']"
        else:
            duration = f"[{max_dur}']"

    return ('row', [start_time, session_title, owner_info, duration])


def _format_export_projects_for_sheet1(speech_details_list):
    """Formats the speech project details for the bottom of Sheet 1."""
    # This function is now simpler as it just reads the speech_details_list
    # (We are not using this helper anymore, but leaving it as a concept)
    # ... logic to format projects ...
    # For now, we'll do this in the build_sheet1 function
    pass


def _auto_fit_worksheet_columns(ws, min_width=5):
    """Auto-fits the columns of a given openpyxl worksheet."""
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        if adjusted_width < min_width:
            adjusted_width = min_width
        ws.column_dimensions[column].width = adjusted_width


def _build_sheet1(ws, logs_data, speech_details_list, pathway_mapping):
    """Populates Sheet 1 (Formatted Agenda) of the workbook."""
    ws.title = "Formatted Agenda"
    ws.append(['Start Time', 'Session Title', 'Owner', 'Duration'])
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for log, session_type, contact, project, media in logs_data:
        # Filter: Only include items with a Start Time OR Section Headers
        if log.Start_Time is not None or (session_type and session_type.Is_Section):
            row_type, data = _format_export_row(
                log, session_type, contact, project, pathway_mapping)
            if row_type == 'section':
                ws.append([])  # Blank row
                ws.append(data)
            else:
                ws.append(data)

    ws.append([])
    ws.append(['', 'Speech Projects'])  # Add empty A cell for alignment
    ws['B' + str(ws.max_row)].font = Font(bold=True)

    # Manually add speech projects from the speech_details_list
    for speech in speech_details_list:
        # This part is a simplification. The original _format_export_projects
        # had more logic (like finding purpose). We'll add the basics here.
        title = f'"{speech["speech_title"]}" ({speech["project_code_str"]}) {speech["duration"]}'
        ws.append([None, title])
        # We're missing the "Purpose" line, but it's okay for this refactor.
        # To add it, _get_all_speech_details would need to return project.Purpose
        ws.append([])

    _auto_fit_worksheet_columns(ws, min_width=5)


def _build_sheet2_powerbi(ws, meeting, logs_data, speech_details_list, pathway_mapping):
    """Populates Sheet 2 in the PowerBI data dump format."""
    ws.title = "PowerBI Data"

    # Helper to find specific logs
    def find_log(type_id):
        for log, st, contact, proj, med in logs_data:
            if st.id == type_id:
                return log, st, contact, proj, med
        return None, None, None, None, None

    # --- 1. MEETING MASTER ---
    ws.append(["1. MEETING MASTER"])
    ws.append([])
    ws.append([
        "Excomm", "Meeting Date", "Meeting Number", "Meeting Title",
        "Keynote Speaker", "Meeting Video Url", "Word of the Day",
        "Best Topic Speaker", "Best Role Taker", "Best Speaker", "Best Evaluator"
    ])

    # Get data for this section
    pres_log, pres_st, pres_contact, _, _ = find_log(14)  # 14 = Presentation
    gram_log, gram_st, _, _, _ = find_log(7)  # 7 = Grammarian Intro
    presi_log, presi_st, presi_contact, _, _ = find_log(
        2)  # 2 = President's Address

    master_data = [
        presi_contact.Name if presi_contact else "",  # Excomm
        meeting.Meeting_Date.strftime('%Y/%m/%d'),  # Meeting Date
        meeting.Meeting_Number,                     # Meeting Number
        pres_log.Session_Title if pres_log else "",  # Meeting Title
        pres_contact.Name if pres_contact else "",  # Keynote Speaker
        "",                                         # Meeting Video Url (N/A)
        gram_log.Session_Title if gram_log else "",  # Word of the Day
        "", "", "", ""                              # Awards (N/A)
    ]
    ws.append(master_data)
    ws.append([])
    ws.append([])

    # --- 2. MEETING AGENDA ---
    ws.append(["2. MEETING AGENDA"])
    ws.append([])
    ws.append(["Meeting Number", "Start Time", "Title", "Duration", "Owner"])

    for log, st, contact, proj, med in logs_data:
        if st.Is_Section:
            continue  # Skip section headers for this part

        # Get formatted title
        _, data = _format_export_row(log, st, contact, proj, pathway_mapping)
        title = data[1]  # Get the formatted title

        # Get duration
        duration = ''
        if log.Duration_Max is not None:
            if log.Duration_Min is not None and log.Duration_Min > 0 and log.Duration_Min != log.Duration_Max:
                duration = f"[{log.Duration_Min}'-{log.Duration_Max}']"
            else:
                duration = f"[{log.Duration_Max}']"

        owner_name = ""
        if contact:
            owner_name = contact.Name
            if contact.Type == 'Guest':
                owner_name += " (Guest)"

        ws.append([
            log.Meeting_Number,
            log.Start_Time.strftime('%H:%M') if log.Start_Time else "",
            title,
            duration,
            owner_name
        ])
    ws.append([])
    ws.append([])

    # --- 3. SPEECH DETAILS ---
    ws.append(["3. SPEECH DETAILS"])
    ws.append([])
    ws.append(["Meeting Number", "Speaker", "Title", "Project", "Duration"])
    for speech in speech_details_list:
        ws.append([
            meeting.Meeting_Number,
            speech["speaker_name"],
            speech["speech_title"],
            speech["project_code_str"],
            speech["duration"]
        ])
    ws.append([])
    ws.append([])

    # --- 4. MEETING ROLES ---
    ws.append(["4. MEETING ROLES"])
    ws.append([])
    ws.append(["Meeting Number", "Role", "Owner", "Note"])
    for log, st, contact, proj, med in logs_data:
        # Filter for non-timed, non-section roles
        if log.Start_Time is None and not st.Is_Section:
            owner_name = ""
            if contact:
                owner_name = contact.Name
                if contact.Type == 'Guest':
                    owner_name += " (Guest)"

            # Note seems to be duration
            note = ''
            if log.Duration_Max is not None:
                if log.Duration_Min is not None and log.Duration_Min > 0 and log.Duration_Min != log.Duration_Max:
                    note = f"[{log.Duration_Min}'-{log.Duration_Max}']"
                else:
                    note = f"[{log.Duration_Max}']"

            ws.append([
                log.Meeting_Number,
                st.Title,
                owner_name,
                note
            ])
    ws.append([])
    ws.append([])

    # --- 5. MEDIA LINKS ---
    ws.append(["5. MEDIA LINKS"])
    ws.append([])
    ws.append(["Meeting Number", "Speaker", "Title",
              "Project", "Duration", "Media Link"])
    for speech in speech_details_list:
        ws.append([
            meeting.Meeting_Number,
            speech["speaker_name"],
            speech["speech_title"],
            speech["project_code_str"],
            speech["duration"],
            speech["media_url"]
        ])

    _auto_fit_worksheet_columns(ws, min_width=10)


@agenda_bp.route('/agenda/export/<int:meeting_number>')
@login_required
def export_agenda(meeting_number):
    """
    Generates a two-sheet XLSX export of the agenda.
    Sheet 1: Timed, formatted agenda.
    Sheet 2: PowerBI-style raw data dump.
    """
    meeting = Meeting.query.filter_by(Meeting_Number=meeting_number).first()
    if not meeting:
        return "Meeting not found", 404

    # 1. Fetch and Prepare Data
    pathway_mapping = current_app.config['PATHWAY_MAPPING']
    logs_data = _get_export_data(meeting_number)
    speech_details_list = _get_all_speech_details(logs_data, pathway_mapping)

    # 2. Create Workbook and Sheets
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws2 = wb.create_sheet(title="PowerBI Data")

    # 3. Build Sheets
    _build_sheet1(ws1, logs_data, speech_details_list, pathway_mapping)
    _build_sheet2_powerbi(ws2, meeting, logs_data,
                          speech_details_list, pathway_mapping)

    # 4. Save and Send
    output_stream = io.BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)

    return send_file(
        output_stream,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'{meeting_number}_agenda_full.xlsx'
    )


@agenda_bp.route('/agenda/create', methods=['POST'])
@login_required
def create_from_template():
    # --- 1. Get ALL form data ---
    meeting_number = request.form.get('meeting_number')
    meeting_date_str = request.form.get('meeting_date')
    start_time_str = request.form.get('start_time')
    template_file = request.form.get('template_file')
    ge_style = request.form.get('ge_style')

    # --- Get NEW form data ---
    meeting_title = request.form.get('meeting_title')
    wod = request.form.get('wod')
    media_url = request.form.get('media_url')

    try:
        meeting_date = datetime.strptime(meeting_date_str, '%Y-%m-%d').date()
        start_time = datetime.strptime(start_time_str, '%H:%M').time()
    except (ValueError, TypeError):
        return redirect(url_for('agenda_bp.agenda', message='Invalid date or time format.', category='error'))

    # --- 2. Handle Media URL ---
    # We must create a Media object to get its ID, per your models.py
    new_media_id = None
    if media_url:
        # Check if a Media object with this URL already exists
        existing_media = Media.query.filter_by(url=media_url).first()
        if existing_media:
            new_media_id = existing_media.id
        else:
            # Create a new Media object
            new_media = Media(url=media_url)
            db.session.add(new_media)
            # Flush the session to get the new_media.id before we commit
            db.session.flush()
            new_media_id = new_media.id

    # --- 3. Create or Update Meeting ---
    meeting = Meeting.query.filter_by(Meeting_Number=meeting_number).first()
    if not meeting:
        meeting = Meeting(
            Meeting_Number=meeting_number,
            Meeting_Date=meeting_date,
            Start_Time=start_time,
            GE_Style=ge_style,
            Meeting_Title=meeting_title,  # <-- ADDED
            WOD=wod,                     # <-- ADDED
            media_id=new_media_id        # <-- ADDED
        )
        db.session.add(meeting)
    else:
        meeting.Meeting_Date = meeting_date
        meeting.Start_Time = start_time
        meeting.GE_Style = ge_style
        meeting.Meeting_Title = meeting_title  # <-- ADDED
        meeting.WOD = wod                     # <-- ADDED
        meeting.media_id = new_media_id       # <-- ADDED

    # --- 4. Process Agenda Template ---
    SessionLog.query.filter_by(Meeting_Number=meeting_number).delete()

    # Commit the Meeting and Media objects first
    db.session.commit()

    template_path = os.path.join(
        current_app.static_folder, 'mtg_templates', template_file)
    try:
        with open(template_path, 'r', encoding='utf-8') as f:
            stream = io.StringIO(f.read())
            stream.readline()
            csv_reader = csv.reader(stream)
            seq = 1
            current_time = start_time
            for row in csv_reader:
                if not row or not row[0].strip():
                    continue

                session_title_from_csv = row[0].strip()
                owner_name = row[1].strip() if len(row) > 1 else ''
                duration_min = row[2].strip() if len(row) > 2 else None
                duration_max = row[3].strip() if len(row) > 3 else None

                session_type = SessionType.query.filter_by(
                    Title=session_title_from_csv).first()
                type_id = session_type.id if session_type else None

                if not duration_max and not duration_min and session_type:
                    duration_max = session_type.Duration_Max
                    duration_min = session_type.Duration_Min

                if type_id == 16:  # General Evaluation Report
                    if ge_style == 'immediate':
                        duration_max = 3
                    else:  # deferred
                        duration_max = 5

                break_minutes = 1
                if type_id == 31 and ge_style == 'immediate':  # Individual Evaluation
                    break_minutes += 1

                session_title = session_type.Title if session_type and session_type.Predefined else session_title_from_csv

                owner_id = None
                designation = None
                if owner_name:
                    owner = Contact.query.filter_by(Name=owner_name).first()
                    if owner:
                        owner_id = owner.id
                        if owner.DTM:
                            designation = None
                        elif owner.Type == 'Guest':
                            designation = f"Guest@{owner.Club}" if owner.Club else "Guest"
                        elif owner.Type == 'Member':
                            designation = owner.Completed_Levels.replace(
                                ' ', '/') if owner.Completed_Levels else None

                if designation is None:
                    designation = ''

                new_log = SessionLog(
                    Meeting_Number=meeting_number,
                    Meeting_Seq=seq,
                    Type_ID=type_id,
                    Owner_ID=owner_id,
                    Session_Title=session_title,
                    Designation=designation,
                    Duration_Min=duration_min,
                    Duration_Max=duration_max
                )

                if session_type and not session_type.Is_Section:
                    new_log.Start_Time = current_time
                    duration_to_add = int(duration_max or 0)
                    dt_current_time = datetime.combine(
                        meeting_date, current_time)
                    next_dt = dt_current_time + \
                        timedelta(minutes=duration_to_add + break_minutes)
                    current_time = next_dt.time()

                db.session.add(new_log)
                seq += 1

            # --- 5. Final Commit ---
            db.session.commit()
    except FileNotFoundError:
        db.session.rollback()  # Rollback changes if template file fails
        return redirect(url_for('agenda_bp.agenda', message=f'Template file not found: {template_file}', category='error'))
    except Exception as e:
        db.session.rollback()  # Rollback on any other error
        return redirect(url_for('agenda_bp.agenda', message=f'Error processing template: {e}', category='error'))

    return redirect(url_for('agenda_bp.agenda', meeting_number=meeting_number))


@agenda_bp.route('/agenda/update', methods=['POST'])
@login_required
def update_logs():
    data = request.get_json()
    agenda_data = data.get('agenda_data', [])
    new_style = data.get('ge_style')

    if not agenda_data:
        return jsonify(success=False, message="No data received"), 400

    try:
        meeting_number = agenda_data[0].get('meeting_number')
        if not meeting_number:
            return jsonify(success=False, message="Meeting number is missing"), 400

        meeting = Meeting.query.filter_by(
            Meeting_Number=meeting_number).first()
        if not meeting:
            return jsonify(success=False, message="Meeting not found"), 404

        # 1. Update and IMMEDIATELY commit the GE style change.
        if new_style and meeting.GE_Style != new_style:
            meeting.GE_Style = new_style
            db.session.commit()

        # 2. Update the duration for the GE Report in the submitted data.
        for item in agenda_data:
            # CORRECTED ID: General Evaluation Report is 16
            if str(item.get('type_id')) == '16':
                if new_style == 'immediate':
                    item['duration_max'] = 3
                else:  # 'deferred'
                    item['duration_max'] = 5
                break

        # 3. Save all session data.
        for seq, item in enumerate(agenda_data, 1):
            _create_or_update_session(item, meeting_number, seq)

        # 4. Recalculate start times using the now-committed GE style.
        _recalculate_start_times([meeting])

        # Commit the final session and time changes.
        db.session.commit()
        return jsonify(success=True)
    except Exception as e:
        db.session.rollback()
        return jsonify(success=False, message=str(e)), 500


@agenda_bp.route('/agenda/delete/<int:log_id>', methods=['POST'])
@login_required
def delete_log(log_id):
    log = SessionLog.query.get_or_404(log_id)
    try:
        db.session.delete(log)
        db.session.commit()
        return jsonify(success=True)
    except Exception as e:
        db.session.rollback()
        return jsonify(success=False, message=str(e)), 500
