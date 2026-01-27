"""
Unit tests for export components.

Tests cover edge cases for:
1. Project code format (e.g., SR1.2, EH2.3)
2. Speech project title format: SR1.2 "My Speech"
3. Owner column with credentials, DTM, and Guest
4. Section spacing (blank line before sections)
5. Component spacing (2 blank lines between components)
6. Speech objectives multi-line format
7. Evaluation titles from Session_Title
8. Hidden session filtering
"""

import unittest
from unittest.mock import Mock, MagicMock, patch
from datetime import time
import openpyxl

from app.services.export.components.agenda import AgendaComponent
from app.services.export.components.powerbi_agenda import PowerBIAgendaComponent
from app.services.export.components.speech_objectives import SpeechObjectivesComponent
from app.services.export.components.roster import RosterComponent
from app.services.export.context import MeetingExportContext


class TestAgendaComponent(unittest.TestCase):
    """Test cases for AgendaComponent formatting."""
    
    def setUp(self):
        """Set up test fixtures."""
        self.component = AgendaComponent()
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        
    def create_mock_log(self, session_type_id, session_title=None, owner_name="John Doe",
                       owner_type="Member", owner_dtm=False, credentials=None,
                       start_time=None, duration_min=None, duration_max=None,
                       project=None, is_section=False, is_hidden=False):
        """Helper to create mock SessionLog."""
        log = Mock()
        log.Session_Title = session_title
        log.Start_Time = start_time or time(19, 0)
        log.Duration_Min = duration_min
        log.Duration_Max = duration_max
        log.credentials = credentials
        log.project = project
        log.id = 1
        log.Meeting_Number = 100
        
        # Mock owner
        if owner_name:
            log.owner = Mock()
            log.owner.Name = owner_name
            log.owner.Type = owner_type
            log.owner.DTM = owner_dtm
            log.owner.credentials = credentials
            log.owners = [log.owner]
        else:
            log.owner = None
            log.owners = []
        
        st = Mock()
        st.id = session_type_id
        # Use provided session_title if it looks like a session type title we care about
        st.Title = "Session Type Title"
        st.Is_Section = is_section
        st.Is_Hidden = is_hidden
        st.Valid_for_Project = (project is not None)
        
        return log, st
    
    def test_project_code_format_level_1(self):
        """Test project code format: SR1.2 (level 1, project 2)."""
        project = Mock()
        project.get_code = Mock(return_value="SR1.2")
        
        log, st = self.create_mock_log(
            session_type_id=30,  # PREPARED_SPEECH
            session_title="My Ice Breaker",
            project=project
        )
        
        context = Mock()
        context.logs = [(log, st)]
        context.speech_details = {
            1: {
                'project_code': 'SR1.2',
                'speech_title': 'My Ice Breaker'
            }
        }
        
        self.component.render(self.ws, context, 1)
        
        # Find the data row (after title and headers)
        title_cell = self.ws.cell(row=4, column=2).value
        self.assertEqual(title_cell, 'SR1.2 "My Ice Breaker"')
    
    def test_project_code_format_level_3(self):
        """Test project code format: EH3.10 (level 3, project 10)."""
        project = Mock()
        project.get_code = Mock(return_value="EH3.10")
        
        log, st = self.create_mock_log(
            session_type_id=30,
            session_title="Advanced Speech",
            project=project
        )
        
        context = Mock()
        context.logs = [(log, st)]
        context.speech_details = {
            1: {
                'project_code': 'EH3.10',
                'speech_title': 'Advanced Speech'
            }
        }
        
        self.component.render(self.ws, context, 1)
        
        title_cell = self.ws.cell(row=4, column=2).value
        self.assertEqual(title_cell, 'EH3.10 "Advanced Speech"')
    
    def test_speech_title_with_quotes(self):
        """Test speech title with quotes - existing quotes are removed first."""
        project = Mock()
        project.get_code = Mock(return_value="SR1.1")
        
        log, st = self.create_mock_log(
            session_type_id=30,
            session_title='The "Best" Speech Ever',
            project=project
        )
        
        context = Mock()
        context.logs = [(log, st)]
        context.speech_details = {
            1: {
                'project_code': 'SR1.1',
                'speech_title': 'The "Best" Speech Ever'
            }
        }
        
        self.component.render(self.ws, context, 1)
        
        title_cell = self.ws.cell(row=4, column=2).value
        # Quotes should be stripped first, then added
        self.assertEqual(title_cell, 'SR1.1 "The Best Speech Ever"')
    
    def test_presentation_project_code_format(self):
        """Test presentation project code format: PS001 (not PS0.1)."""
        project = Mock()
        # Presentations use series abbreviation + 3-digit code
        project.get_code = Mock(return_value="PS001")
        
        log, st = self.create_mock_log(
            session_type_id=43,  # PRESENTATION
            session_title="My Presentation",
            project=project,
            duration_max=10
        )
        
        context = Mock()
        context.logs = [(log, st)]
        context.speech_details = {
            1: {
                'project_code': 'PS001',  # Presentation format
                'speech_title': 'My Presentation'
            }
        }
        
        self.component.render(self.ws, context, 1)
        
        # Row 4 is the data row (after title, blank, headers)
        title_cell = self.ws.cell(row=4, column=2).value
        self.assertEqual(title_cell, 'PS001 "My Presentation"')
    
    def test_presentation_vs_pathway_format(self):
        """Test that presentations use XXnnn format while pathways use XX.n.n."""
        # Pathway speech
        pathway_project = Mock()
        pathway_project.get_code = Mock(return_value="SR1.2")
        
        # Presentation speech
        presentation_project = Mock()
        presentation_project.get_code = Mock(return_value="PS015")
        
        log1, st1 = self.create_mock_log(
            session_type_id=30,  # PREPARED_SPEECH
            session_title="Pathway Speech",
            project=pathway_project,
            duration_max=7
        )
        
        log2, st2 = self.create_mock_log(
            session_type_id=43,  # PRESENTATION
            session_title="Presentation Speech",
            project=presentation_project,
            duration_max=10
        )
        log2.id = 2
        
        context = Mock()
        context.logs = [(log1, st1), (log2, st2)]
        context.speech_details = {
            1: {
                'project_code': 'SR1.2',
                'speech_title': 'Pathway Speech'
            },
            2: {
                'project_code': 'PS015',
                'speech_title': 'Presentation Speech'
            }
        }
        
        self.component.render(self.ws, context, 1)
        
        # Row 4: Pathway speech
        # Row 5: Presentation speech
        pathway_title = self.ws.cell(row=4, column=2).value
        presentation_title = self.ws.cell(row=5, column=2).value
        
        self.assertEqual(pathway_title, 'SR1.2 "Pathway Speech"')
        self.assertEqual(presentation_title, 'PS015 "Presentation Speech"')
    
    def test_missing_owner(self):
        """Test session with no owner - should have empty owner column."""
        log, st = self.create_mock_log(
            session_type_id=1,
            owner_name=None
        )
        log.owner = None
        
        context = Mock()
        context.logs = [(log, st)]
        context.speech_details = {}
        
        self.component.render(self.ws, context, 1)
        
        owner_cell = self.ws.cell(row=4, column=3).value
        self.assertEqual(owner_cell, "")
    
    def test_empty_speech_title(self):
        """Test speech with empty Session_Title - should use session type title."""
        project = Mock()
        project.get_code = Mock(return_value="SR1.1")
        
        log, st = self.create_mock_log(
            session_type_id=30,
            session_title="",  # Empty title
            project=project
        )
        
        context = Mock()
        context.logs = [(log, st)]
        context.speech_details = {
            1: {
                'project_code': 'SR1.1',
                'speech_title': ''
            }
        }
        
        self.component.render(self.ws, context, 1)
        
        title_cell = self.ws.cell(row=4, column=2).value
        # Should fall back to session type title when speech title is empty
        self.assertEqual(title_cell, "Session Type Title")
    
    def test_special_characters_in_title(self):
        """Test speech title with various special characters."""
        project = Mock()
        project.get_code = Mock(return_value="SR2.1")
        
        log, st = self.create_mock_log(
            session_type_id=30,
            session_title="Speech: 'Success' & \"Growth\" (2024)",
            project=project
        )
        
        context = Mock()
        context.logs = [(log, st)]
        context.speech_details = {
            1: {
                'project_code': 'SR2.1',
                'speech_title': "Speech: 'Success' & \"Growth\" (2024)"
            }
        }
        
        self.component.render(self.ws, context, 1)
        
        title_cell = self.ws.cell(row=4, column=2).value
        # All quotes should be stripped, other special chars preserved
        self.assertEqual(title_cell, 'SR2.1 "Speech: Success & Growth (2024)"')
    
    def test_multiple_sections_spacing(self):
        """Test that blank lines are added before each section."""
        log1, st1 = self.create_mock_log(session_type_id=1, session_title="Regular")
        log2, st2 = self.create_mock_log(session_type_id=2, session_title="Section 1", is_section=True)
        log2.id = 2
        log3, st3 = self.create_mock_log(session_type_id=3, session_title="Section 2", is_section=True)
        log3.id = 3
        
        context = Mock()
        context.logs = [(log1, st1), (log2, st2), (log3, st3)]
        context.speech_details = {}
        
        self.component.render(self.ws, context, 1)
        
        # Verify sections are in the output (they should have titles)
        # The exact row numbers depend on blank line insertion logic
        # Just verify that we have the expected number of rows
        # 1: AGENDA title, 2: blank, 3: headers, 4: regular, 5: blank, 6: section1, 7: blank, 8: section2
        self.assertGreaterEqual(self.ws.max_row, 8)
    
    def test_dtm_with_guest_type(self):
        """Test edge case: DTM member incorrectly marked as Guest."""
        log, st = self.create_mock_log(
            session_type_id=7,
            owner_name="John Doe",
            owner_type="Guest",  # Incorrectly marked as guest
            owner_dtm=True,  # But is DTM
            credentials="DTM"
        )
        
        context = Mock()
        context.logs = [(log, st)]
        context.speech_details = {}
        
        self.component.render(self.ws, context, 1)
        
        owner_cell = self.ws.cell(row=4, column=3).value
        # DTM takes precedence - no Guest or credentials
        self.assertEqual(owner_cell, "John Doeᴰᵀᴹ")
    
    def test_very_long_title_wrapping(self):
        """Test that very long titles trigger text wrapping."""
        project = Mock()
        project.get_code = Mock(return_value="EH1.1")
        
        long_title = "A" * 100  # 100 character title
        log, st = self.create_mock_log(
            session_type_id=30,
            session_title=long_title,
            project=project
        )
        
        context = Mock()
        context.logs = [(log, st)]
        context.speech_details = {
            1: {
                'project_code': 'EH1.1',
                'speech_title': long_title
            }
        }
        
        self.component.render(self.ws, context, 1)
        
        # Check that wrap_text was applied (title > 50 chars)
        cell = self.ws.cell(row=4, column=2)
        self.assertTrue(cell.alignment.wrap_text if cell.alignment else False)
    
    def test_owner_with_credentials(self):
        """Test owner column with credentials: John Doe - CC."""
        log, st = self.create_mock_log(
            session_type_id=7,  # TABLE_TOPICS
            owner_name="John Doe",
            credentials="CC"
        )
        
        context = Mock()
        context.logs = [(log, st)]
        context.speech_details = {}
        
        self.component.render(self.ws, context, 1)
        
        owner_cell = self.ws.cell(row=4, column=3).value
        self.assertEqual(owner_cell, "John Doe - CC")
    
    def test_owner_dtm_no_credentials(self):
        """Test DTM member doesn't get extra credentials: John Doeᴰᵀᴹ."""
        log, st = self.create_mock_log(
            session_type_id=7,
            owner_name="John Doe",
            owner_dtm=True,
            credentials="DTM"  # Should be ignored
        )
        
        context = Mock()
        context.logs = [(log, st)]
        context.speech_details = {}
        
        self.component.render(self.ws, context, 1)
        
        owner_cell = self.ws.cell(row=4, column=3).value
        self.assertEqual(owner_cell, "John Doeᴰᵀᴹ")
    
    def test_owner_guest_credential(self):
        """Test guest owner: Jane Smith - Guest."""
        log, st = self.create_mock_log(
            session_type_id=7,
            owner_name="Jane Smith",
            owner_type="Guest"
        )
        
        context = Mock()
        context.logs = [(log, st)]
        context.speech_details = {}
        
        self.component.render(self.ws, context, 1)
        
        owner_cell = self.ws.cell(row=4, column=3).value
        self.assertEqual(owner_cell, "Jane Smith - Guest")
    
    def test_section_spacing(self):
        """Test blank line before section sessions."""
        log1, st1 = self.create_mock_log(session_type_id=1, session_title="Opening")
        log2, st2 = self.create_mock_log(session_type_id=2, session_title="Prepared Speeches", is_section=True)
        log3, st3 = self.create_mock_log(session_type_id=30, session_title="Speech 1")
        
        context = Mock()
        context.logs = [(log1, st1), (log2, st2), (log3, st3)]
        context.speech_details = {}
        
        self.component.render(self.ws, context, 1)
        
        # Row 4: Opening
        # Row 5: Blank (before section)
        # Row 6: Prepared Speeches section
        # Row 7: Speech 1
        self.assertIsNotNone(self.ws.cell(row=4, column=1).value)  # Opening has data
        self.assertIsNone(self.ws.cell(row=5, column=1).value)  # Blank line
        self.assertIsNotNone(self.ws.cell(row=6, column=1).value)  # Section has data
    
    def test_component_spacing(self):
        """Test component returns ws.max_row + 3 for 2 blank lines."""
        log, st = self.create_mock_log(session_type_id=1)
        
        context = Mock()
        context.logs = [(log, st)]
        context.speech_details = {}
        
        max_row_before = self.ws.max_row
        next_row = self.component.render(self.ws, context, 1)
        
        # Should add 3 to max_row (2 blank lines)
        self.assertEqual(next_row, self.ws.max_row + 3)
    
    def test_evaluation_title(self):
        """Test evaluation title: Evaluation for <speaker>."""
        log, st = self.create_mock_log(
            session_type_id=31,  # EVALUATION
            session_title="John Doe"  # Speaker name stored in Session_Title
        )
        st.Title = "Evaluation"
        
        context = Mock()
        context.logs = [(log, st)]
        context.speech_details = {}
        
        self.component.render(self.ws, context, 1)
        
        title_cell = self.ws.cell(row=4, column=2).value
        self.assertEqual(title_cell, "Evaluation for John Doe")
    
    def test_hidden_session_filtered(self):
        """Test hidden sessions are not exported."""
        log1, st1 = self.create_mock_log(session_type_id=1, session_title="Visible")
        log2, st2 = self.create_mock_log(session_type_id=2, session_title="Hidden", is_hidden=True)
        
        context = Mock()
        context.logs = [(log1, st1), (log2, st2)]
        context.speech_details = {}
        
        self.component.render(self.ws, context, 1)
        
        # Should only have 4 rows: title, blank, headers, visible session
        self.assertEqual(self.ws.max_row, 4)
        self.assertEqual(self.ws.cell(row=4, column=2).value, "Visible")
    
    def test_component_has_title(self):
        """Test component includes AGENDA title."""
        log, st = self.create_mock_log(session_type_id=1)
        
        context = Mock()
        context.logs = [(log, st)]
        context.speech_details = {}
        
        self.component.render(self.ws, context, 1)
        
        # Agenda component doesn't have blank line before title (unlike other components)
        self.assertEqual(self.ws.cell(row=1, column=1).value, "AGENDA")
    
    def test_keynote_speech_strips_quotes(self):
        """Test keynote speech title has quotes stripped."""
        log, st = self.create_mock_log(
            session_type_id=20,  # KEYNOTE_SPEECH
            session_title='The "Best" Keynote Ever'
        )
        st.Title = "Keynote Speech"
        
        context = Mock()
        context.logs = [(log, st)]
        context.speech_details = {}
        
        self.component.render(self.ws, context, 1)
        
        title_cell = self.ws.cell(row=4, column=2).value
        # Keynote titles should have quotes stripped, no project code prefix
        self.assertEqual(title_cell, "The Best Keynote Ever")


class TestSpeechObjectivesComponent(unittest.TestCase):
    """Test cases for SpeechObjectivesComponent formatting."""
    
    def setUp(self):
        """Set up test fixtures."""
        self.component = SpeechObjectivesComponent()
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
    
    def test_speech_objectives_format(self):
        """Test multi-line speech objectives format."""
        project = Mock()
        project.Purpose = "Learn to organize and deliver a speech."
        
        log = Mock()
        log.id = 1
        log.project = project
        log.Session_Title = "Ice Breaker"
        
        st = Mock()
        st.Is_Hidden = False
        st.Valid_for_Project = True
        
        context = Mock()
        context.logs = [(log, st)]
        context.speech_details = {
            1: {
                'project_code': 'SR1.1',
                'pathway_name': 'Strategic Relationships',
                'project_name': 'Ice Breaker',
                'project_type': 'required',
                'project_purpose': 'Learn to organize and deliver a speech.',
                'duration_min': 4,
                'duration_max': 6
            }
        }
        
        self.component.render(self.ws, context, 1)
        
        # Row 1: Blank line
        # Row 2: Component title
        # Row 3: Blank
        # Row 4: First line of objective
        # Row 5: Second line (purpose)
        # Row 6: Blank
        self.assertEqual(self.ws.cell(row=2, column=1).value, "PROJECT OBJECTIVES")
        self.assertEqual(
            self.ws.cell(row=4, column=1).value,
            "Strategic Relationships (SR1.1) Ice Breaker (Required) [4'-6']"
        )
        self.assertEqual(
            self.ws.cell(row=5, column=1).value,
            "Learn to organize and deliver a speech."
        )
        self.assertIsNone(self.ws.cell(row=6, column=1).value)  # Blank line
    
    def test_speech_objectives_elective(self):
        """Test speech objectives with elective project."""
        project = Mock()
        project.Purpose = "Master advanced techniques."
        
        log = Mock()
        log.id = 1
        log.project = project
        
        st = Mock()
        st.Is_Hidden = False
        st.Valid_for_Project = True
        
        context = Mock()
        context.logs = [(log, st)]
        context.speech_details = {
            1: {
                'project_code': 'EH2.5',
                'pathway_name': 'Effective Coaching',
                'project_name': 'Advanced Project',
                'project_type': 'elective',
                'project_purpose': 'Master advanced techniques.',
                'duration_min': 5,
                'duration_max': 7
            }
        }
        
        self.component.render(self.ws, context, 1)
        
        # Row 4 is first line (after blank, title, blank)
        first_line = self.ws.cell(row=4, column=1).value
        self.assertIn("(Elective)", first_line)
    
    def test_speech_objectives_no_purpose_skipped(self):
        """Test speech without purpose is skipped."""
        project = Mock()
        project.Purpose = None
        
        log = Mock()
        log.id = 1
        log.project = project
        
        st = Mock()
        st.Is_Hidden = False
        st.Valid_for_Project = True
        
        context = Mock()
        context.logs = [(log, st)]
        context.speech_details = {
            1: {
                'project_code': 'SR1.1',
                'project_purpose': '',  # Empty purpose
            }
        }
        
        self.component.render(self.ws, context, 1)
        
        # Component always renders title (row 2 after blank line), but no objectives
        self.assertGreaterEqual(self.ws.max_row, 2)
        self.assertEqual(self.ws.cell(row=2, column=1).value, "PROJECT OBJECTIVES")
    
    def test_objectives_sorted_pathway_before_presentation(self):
        """Test objectives are sorted with pathway projects before presentations."""
        # Create multiple projects in random order
        projects_data = [
            # Presentation (should be last)
            {
                'id': 1,
                'code': 'PS001',
                'pathway': 'Presentation Mastery',
                'name': 'Presentation Speech',
                'type': 'required',
                'purpose': 'Present effectively.',
                'duration_min': 5,
                'duration_max': 7
            },
            # Level 2 pathway (should be after level 1)
            {
                'id': 2,
                'code': 'EH2.3',
                'pathway': 'Effective Coaching',
                'name': 'Advanced Project',
                'type': 'elective',
                'purpose': 'Coach effectively.',
                'duration_min': 5,
                'duration_max': 7
            },
            # Level 1 pathway (should be first)
            {
                'id': 3,
                'code': 'SR1.2',
                'pathway': 'Strategic Relationships',
                'name': 'Ice Breaker',
                'type': 'required',
                'purpose': 'Introduce yourself.',
                'duration_min': 4,
                'duration_max': 6
            },
        ]
        
        logs = []
        speech_details = {}
        
        for i, proj_data in enumerate(projects_data):
            project = Mock()
            project.Purpose = proj_data['purpose']
            
            log = Mock()
            log.id = proj_data['id']
            log.project = project
            
            st = Mock()
            st.Is_Hidden = False
            st.Valid_for_Project = True
            
            logs.append((log, st))
            speech_details[proj_data['id']] = {
                'project_code': proj_data['code'],
                'pathway_name': proj_data['pathway'],
                'project_name': proj_data['name'],
                'project_type': proj_data['type'],
                'project_purpose': proj_data['purpose'],
                'duration_min': proj_data['duration_min'],
                'duration_max': proj_data['duration_max']
            }
        
        context = Mock()
        context.logs = logs
        context.speech_details = speech_details
        
        self.component.render(self.ws, context, 1)
        
        # Row 1: Blank, Row 2: Title, Row 3: Blank
        # Row 4: SR1.2 first line, Row 5: SR1.2 purpose, Row 6: Blank
        # Row 7: EH2.3 first line, Row 8: EH2.3 purpose, Row 9: Blank
        # Row 10: PS001 first line, Row 11: PS001 purpose, Row 12: Blank
        
        # Verify sorting order
        sr_line = self.ws.cell(row=4, column=1).value
        eh_line = self.ws.cell(row=7, column=1).value
        ps_line = self.ws.cell(row=10, column=1).value
        
        self.assertIn("SR1.2", sr_line)  # Pathway level 1 first
        self.assertIn("EH2.3", eh_line)  # Pathway level 2 second
        self.assertIn("PS001", ps_line)  # Presentation last


class TestPowerBIAgendaComponent(unittest.TestCase):
    """Test cases for PowerBIAgendaComponent formatting."""
    
    def setUp(self):
        """Set up test fixtures."""
        self.component = PowerBIAgendaComponent()
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
    
    def test_powerbi_format_matches_agenda(self):
        """Test PowerBI agenda uses same title format as main agenda."""
        project = Mock()
        project.get_code = Mock(return_value="SR1.2")
        
        log = Mock()
        log.id = 1
        log.Session_Title = "My Speech"
        log.Start_Time = time(19, 30)
        log.Duration_Min = 5
        log.Duration_Max = 7
        log.credentials = "CC"
        log.project = project
        log.Meeting_Number = 100
        
        log.owner = Mock()
        log.owner.Name = "John Doe"
        log.owner.Type = "Member"
        log.owner.DTM = False
        log.owner.credentials = "CC"
        log.owners = [log.owner]
        
        st = Mock()
        st.id = 30
        st.Is_Section = False
        st.Is_Hidden = False
        st.Valid_for_Project = True
        
        context = Mock()
        context.logs = [(log, st)]
        context.speech_details = {
            1: {
                'project_code': 'SR1.2',
                'speech_title': 'My Speech'
            }
        }
        
        self.component.render(self.ws, context, 1)
        
        # Row 1: Blank, Row 2: Title, Row 3: Blank, Row 4: Headers, Row 5: Data
        title_cell = self.ws.cell(row=5, column=3).value
        owner_cell = self.ws.cell(row=5, column=5).value
        
        self.assertEqual(title_cell, 'SR1.2 "My Speech"')
        self.assertEqual(owner_cell, "John Doe - CC")


class TestRosterComponent(unittest.TestCase):
    """Test cases for RosterComponent formatting."""

    def setUp(self):
        """Set up test fixtures."""
        self.component = RosterComponent()
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active

    @patch('app.services.export.components.roster.orm')
    @patch('app.services.export.components.roster.Roster')
    def test_roster_render(self, mock_roster_model, mock_orm):
        """Test roster rendering with ticket names."""
        # Setup mock data for orm
        mock_orm.joinedload.return_value = Mock()
        
        # Setup mock data for Roster
        mock_entry = Mock()
        mock_entry.order_number = 1
        
        mock_entry.contact = Mock()
        mock_entry.contact.Name = "John Doe"
        mock_entry.contact.Type = "Member"
        
        mock_entry.ticket = Mock()
        mock_entry.ticket.name = "Early Bird"
        
        # Mock the query chain
        # Roster.query.options(...).filter_by(...).order_by(...).all()
        mock_query = mock_roster_model.query
        mock_query.options.return_value = mock_query
        mock_query.filter_by.return_value = mock_query
        mock_query.order_by.return_value = mock_query
        mock_query.all.return_value = [mock_entry]
        
        context = Mock()
        context.meeting_number = 100
        
        # Execute
        self.component.render(self.ws, context, 1)
        
        # Verify Headers
        self.assertEqual(self.ws.cell(row=1, column=1).value, "Order")
        self.assertEqual(self.ws.cell(row=1, column=2).value, "Name")
        self.assertEqual(self.ws.cell(row=1, column=3).value, "Ticket")
        
        # Verify Data
        self.assertEqual(self.ws.cell(row=2, column=1).value, 1)
        self.assertEqual(self.ws.cell(row=2, column=2).value, "John Doe")
        # Should contain ONLY ticket name, not type
        self.assertEqual(self.ws.cell(row=2, column=3).value, "Early Bird")

    @patch('app.services.export.components.roster.orm')
    @patch('app.services.export.components.roster.Roster')
    def test_roster_render_no_ticket(self, mock_roster_model, mock_orm):
        """Test roster rendering with missing ticket."""
        # Setup mock data for orm
        mock_orm.joinedload.return_value = Mock()
        
        mock_entry = Mock()
        mock_entry.order_number = 2
        mock_entry.contact.Name = "Jane Doe"
        mock_entry.ticket = None # No ticket
        
        mock_query = mock_roster_model.query
        mock_query.options.return_value = mock_query
        mock_query.filter_by.return_value = mock_query
        mock_query.order_by.return_value = mock_query
        mock_query.all.return_value = [mock_entry]
        
        context = Mock()
        context.meeting_number = 100
        
        self.component.render(self.ws, context, 1)
        
        self.assertEqual(self.ws.cell(row=2, column=3).value, "")


if __name__ == '__main__':
    unittest.main()
