import os
import sys
import re
import openpyxl
from datetime import datetime

# Add the app directory to the path so we can import from 'app'
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from app import create_app, db
from app.models.meeting import Meeting
from app.models.contact import Contact
from app.models.contact_club import ContactClub
from app.models.session import SessionLog, SessionType, OwnerMeetingRoles
from app.models.media import Media
from app.models.roster import MeetingRole

# Normalization Regex: Matches (PM2.1), DL.1.4.2, (Guest), etc.
NORM_PATTERN = re.compile(r'\s*\(?[A-Z]{2,}[\.\w]*\d+[\.\w]*\)?|\s*\(Guest\)', re.IGNORECASE)

def normalize_name(name):
    if not name:
        return ""
    name_str = str(name).strip()
    return NORM_PATTERN.sub('', name_str).strip()

def get_or_create_contact(name, club_id=None):
    if not name:
        return None, False
    
    # Try normalization
    norm_name = normalize_name(name)
    if not norm_name:
        return None, False
        
    contact = Contact.query.filter(Contact.Name == norm_name).first()
    created = False
    if not contact:
        # Create new contact
        contact = Contact(Name=norm_name, Type='Guest')
        db.session.add(contact)
        db.session.flush()
        created = True
    
    # Ensure association with club
    if club_id:
        existing_assoc = ContactClub.query.filter_by(contact_id=contact.id, club_id=club_id).first()
        if not existing_assoc:
            db.session.add(ContactClub(contact_id=contact.id, club_id=club_id))
            db.session.flush()
            
    return contact, created

def process_yoodli_import(file_path, dry_run=False):
    app = create_app()
    with app.app_context():
        print(f"Opening {file_path}...")
        try:
            wb = openpyxl.load_workbook(file_path)
        except Exception as e:
            print(f"Error opening workbook: {e}")
            return
            
        sheet = wb.active
        
        # Determine column mapping
        headers = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]
        print(f"Detected Headers: {headers}")
        
        col_mtg = 0
        col_speaker = 1
        col_speech_url = 2
        col_evaluator = 3
        col_eval_url = 4
        
        # Adjust based on detected headers or column count
        if len(headers) > 5 and 'Video Link' in headers:
            # Matches the 9-column format seen in inspection
            col_mtg = 0
            col_speaker = 1
            col_speech_url = 5
            col_evaluator = 6
            col_eval_url = 7
            print("Using 9-column mapping.")
        else:
            print("Using default 5-column mapping.")

        results = []
        # Header for results logging
        results_headers = ["Row", "Meeting", "Speaker", "Evaluator", "Status", "Details"]
        
        # Get Individual Evaluator Role
        eval_role = MeetingRole.query.filter_by(name='Individual Evaluator').first()
        if not eval_role:
             print("Critical Error: 'Individual Evaluator' role not found in database.")
             return

        # Get Individual Evaluator SessionType
        eval_st = SessionType.query.filter_by(Title='Evaluation').first()
        if not eval_st:
             # Fallback: find any session type with the eval role
             eval_st = SessionType.query.filter_by(role_id=eval_role.id).first()
             
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue
                
            mtg_num = row[col_mtg]
            speaker_raw = row[col_speaker]
            speech_url = row[col_speech_url]
            evaluator_raw = row[col_evaluator]
            eval_url = row[col_eval_url]
            
            if not mtg_num:
                continue
                
            status = "Success"
            details = []
            
            # 1. Find Meeting
            meeting = Meeting.query.filter_by(Meeting_Number=mtg_num).first()
            if not meeting:
                results.append([row_idx, mtg_num, speaker_raw, evaluator_raw, "Error", "Meeting not found"])
                continue
            
            club_id = meeting.club_id
                
            # 2. Find/Create Contacts and ensure Club Association
            speaker, s_created = get_or_create_contact(speaker_raw, club_id=club_id)
            if s_created: details.append(f"Created speaker contact: {speaker.Name}")
            
            evaluator, e_created = get_or_create_contact(evaluator_raw, club_id=club_id)
            if e_created: details.append(f"Created evaluator contact: {evaluator.Name}")
            
            if not speaker or not evaluator:
                results.append([row_idx, mtg_num, speaker_raw, evaluator_raw, "Error", "Invalid speaker or evaluator name"])
                continue

            # 3. Update Speaker Session
            # Find a session log where the speaker is an owner
            speaker_name_norm = speaker.Name
            
            # Query SessionLogs for this meeting that belong to a "speaker" category role
            speaker_session = db.session.query(SessionLog).join(OwnerMeetingRoles).join(MeetingRole).filter(
                SessionLog.Meeting_Number == mtg_num,
                OwnerMeetingRoles.contact_id == speaker.id,
                MeetingRole.award_category == 'speaker'
            ).first()
            
            if speaker_session:
                if speech_url and str(speech_url).lower() != 'na':
                    speaker_session.update_media(speech_url)
                    details.append("Updated speaker media URL")
            else:
                details.append("Warn: No speaker session found for this contact in meeting")

            # 4. Update Evaluator Session
            # Logic: Match by Session_Title (norm speaker name) -> Match by Blank Title -> Create
            eval_sessions = SessionLog.query.join(SessionType).filter(
                SessionLog.Meeting_Number == mtg_num,
                SessionType.role_id == eval_role.id
            ).all()
            
            target_eval_session = None
            
            # Try to match by title
            for s in eval_sessions:
                if s.Session_Title == speaker_name_norm:
                    target_eval_session = s
                    break
            
            # Try to match by blank title
            if not target_eval_session:
                for s in eval_sessions:
                    if not s.Session_Title or str(s.Session_Title).strip() == "":
                        target_eval_session = s
                        s.Session_Title = speaker_name_norm
                        details.append(f"Assigned blank evaluation session to {speaker_name_norm}")
                        break
            
            # Create new if still not found
            if not target_eval_session:
                if not eval_st:
                     details.append("Error: Cannot create evaluation session - SessionType not found")
                else:
                    # Find max seq
                    from sqlalchemy import func
                    max_seq = db.session.query(func.max(SessionLog.Meeting_Seq)).filter_by(Meeting_Number=mtg_num).scalar() or 0
                    
                    new_eval = SessionLog(
                        Meeting_Number=mtg_num,
                        Meeting_Seq=max_seq + 1,
                        Session_Title=speaker_name_norm,
                        Type_ID=eval_st.id,
                        Status='active'
                    )
                    db.session.add(new_eval)
                    db.session.flush()
                    
                    # Create ownership
                    omr = OwnerMeetingRoles(
                        meeting_id=meeting.id,
                        role_id=eval_role.id,
                        contact_id=evaluator.id,
                        session_log_id=new_eval.id
                    )
                    db.session.add(omr)
                    target_eval_session = new_eval
                    details.append(f"Created new evaluation session for {speaker_name_norm}")
            else:
                # Ensure evaluator is assigned to this session if it was found
                # Check if this evaluator is already an owner, if not, update it
                # For Individual Evaluator, it's usually has_single_owner=True
                existing_omr = OwnerMeetingRoles.query.filter_by(
                    session_log_id=target_eval_session.id,
                    role_id=eval_role.id
                ).first()
                
                if existing_omr:
                    if existing_omr.contact_id != evaluator.id:
                        old_eval_name = existing_omr.contact.Name if existing_omr.contact else "Unknown"
                        existing_omr.contact_id = evaluator.id
                        details.append(f"Reassigned evaluator from {old_eval_name} to {evaluator.Name}")
                else:
                    # Create ownership if missing
                    new_omr = OwnerMeetingRoles(
                        meeting_id=meeting.id,
                        role_id=eval_role.id,
                        contact_id=evaluator.id,
                        session_log_id=target_eval_session.id
                    )
                    db.session.add(new_omr)
                    details.append(f"Assigned evaluator {evaluator.Name}")

            # Update Media for Evaluator
            if target_eval_session and eval_url and str(eval_url).lower() != 'na':
                target_eval_session.update_media(eval_url)
                details.append("Updated evaluator media URL")
                
            results.append([row_idx, mtg_num, speaker_name_norm, evaluator.Name, status, "; ".join(details)])

        if not dry_run:
            db.session.commit()
            print("Changes committed to database.")
        else:
            db.session.rollback()
            print("Dry run: Changes rolled back.")

        # Save Results to Excel
        output_path = "instance/unmatched_yoodli.xlsx"
        out_wb = openpyxl.Workbook()
        out_ws = out_wb.active
        out_ws.append(results_headers)
        for row in results:
            out_ws.append(row)
        
        try:
            out_wb.save(output_path)
            print(f"Details saved to {output_path}")
        except Exception as e:
            print(f"Error saving output: {e}")

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description='Import Yoodli media URLs.')
    parser.add_argument('--file', default='instance/yoodli.xlsx', help='Path to Yoodli Excel file')
    parser.add_argument('--dry-run', action='store_true', help='Perform a dry run without committing changes')
    args = parser.parse_args()
    
    # Resolve absolute path relative to script location
    base_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
    full_path = os.path.join(base_dir, args.file)
    
    if not os.path.exists(full_path):
        print(f"Error: {args.file} not found at {full_path}")
    else:
        process_yoodli_import(full_path, dry_run=args.dry_run)
